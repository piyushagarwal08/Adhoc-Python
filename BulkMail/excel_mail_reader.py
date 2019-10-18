import openpyxl
workbook = openpyxl.load_workbook("mails.xlsx")

# INDIA PAPER MILLS
sheet1 = workbook.get_sheet_by_name("Sheet2")

row = sheet1.max_row
print(f"total number of entries: {row}") # total no of entries

for i in range(1,row):
    print(sheet1.cell(row=i,column=1).value)

